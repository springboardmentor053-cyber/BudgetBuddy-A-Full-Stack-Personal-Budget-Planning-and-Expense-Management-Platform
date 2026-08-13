import calendar
import io
import json
from datetime import datetime, date
from decimal import Decimal
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.pdfgen import canvas

from expenses.models import Expense, Income, Budget
from savings.models import SavingsGoal, Notification


def parse_date_range(request):
    """
    Parses date range filters from request query parameters.
    """
    today = timezone.now().date()
    period = request.query_params.get('period', '').lower()
    month_param = request.query_params.get('month')
    year_param = request.query_params.get('year')
    start_date_param = request.query_params.get('start_date')
    end_date_param = request.query_params.get('end_date')

    if start_date_param or end_date_param:
        try:
            s_date = datetime.strptime(start_date_param, '%Y-%m-%d').date() if start_date_param else date(2000, 1, 1)
        except ValueError:
            s_date = date(2000, 1, 1)

        try:
            e_date = datetime.strptime(end_date_param, '%Y-%m-%d').date() if end_date_param else date(2099, 12, 31)
        except ValueError:
            e_date = date(2099, 12, 31)

        is_invalid = s_date > e_date

        return {
            'start_date': s_date,
            'end_date': e_date,
            'is_invalid': is_invalid,
            'month': s_date.month,
            'year': s_date.year,
            'period_label': f"Custom ({s_date.isoformat()} to {e_date.isoformat()})"
        }


    if period == 'previous_month':
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year

        _, last_day = calendar.monthrange(prev_year, prev_month)
        s_date = date(prev_year, prev_month, 1)
        e_date = date(prev_year, prev_month, last_day)

        return {
            'start_date': s_date,
            'end_date': e_date,
            'month': prev_month,
            'year': prev_year,
            'period_label': f"Previous Month ({prev_year}-{prev_month:02d})"
        }

    if month_param and year_param:
        try:
            m = int(month_param)
            y = int(year_param)
            _, last_day = calendar.monthrange(y, m)
            s_date = date(y, m, 1)
            e_date = date(y, m, last_day)
            return {
                'start_date': s_date,
                'end_date': e_date,
                'month': m,
                'year': y,
                'period_label': f"{y}-{m:02d}"
            }
        except (ValueError, TypeError):
            pass

    curr_month = today.month
    curr_year = today.year
    _, last_day = calendar.monthrange(curr_year, curr_month)
    s_date = date(curr_year, curr_month, 1)
    e_date = date(curr_year, curr_month, last_day)

    return {
        'start_date': s_date,
        'end_date': e_date,
        'month': curr_month,
        'year': curr_year,
        'period_label': f"Current Month ({curr_year}-{curr_month:02d})"
    }


class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to dynamically compute total pages and draw production footers.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor('#64748B'))

        # Header line & branding on pages > 1
        if self._pageNumber > 1:
            self.drawString(36, 755, "BudgetBuddy — Financial Summary Report")
            self.drawRightString(612 - 36, 755, timezone.now().strftime('%Y-%m-%d'))
            self.setStrokeColor(colors.HexColor('#E2E8F0'))
            self.setLineWidth(0.5)
            self.line(36, 748, 612 - 36, 748)

        # Footer
        footer_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(612 - 36, 20, footer_text)
        timestamp_str = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Confidential Financial Statement"
        self.drawString(36, 20, timestamp_str)

        self.setStrokeColor(colors.HexColor('#E2E8F0'))
        self.setLineWidth(0.5)
        self.line(36, 32, 612 - 36, 32)

        self.restoreState()


class ReportExportHandler:
    """
    Production-quality Handler for generating PDF, Excel, and JSON exports according to Module 9 specifications.
    """

    @staticmethod
    def export_json(data):
        return json.dumps(data, indent=2, default=str)

    @staticmethod
    def export_pdf(user, report_data):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=48,
            bottomMargin=48
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#1E3A8A'),
            fontName='Helvetica-Bold'
        )

        subtitle_style = ParagraphStyle(
            'DocSubtitle',
            parent=styles['Normal'],
            fontSize=9.5,
            leading=14,
            textColor=colors.HexColor('#64748B'),
            fontName='Helvetica'
        )

        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=13,
            leading=17,
            textColor=colors.HexColor('#1E293B'),
            fontName='Helvetica-Bold',
            spaceBefore=14,
            spaceAfter=8
        )

        cell_left = ParagraphStyle('CellLeft', fontSize=8.5, leading=11, fontName='Helvetica', alignment=TA_LEFT)
        cell_right = ParagraphStyle('CellRight', fontSize=8.5, leading=11, fontName='Helvetica', alignment=TA_RIGHT)
        cell_bold_left = ParagraphStyle('CellBoldLeft', fontSize=8.5, leading=11, fontName='Helvetica-Bold', alignment=TA_LEFT)
        cell_bold_right = ParagraphStyle('CellBoldRight', fontSize=8.5, leading=11, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        cell_hdr = ParagraphStyle('CellHdr', fontSize=9, leading=11, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_LEFT)
        cell_hdr_right = ParagraphStyle('CellHdrRight', fontSize=9, leading=11, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_RIGHT)

        elements = []

        # Document Header Banner
        elements.append(Paragraph("BudgetBuddy — Financial Summary Report", title_style))
        period_str = report_data.get('financial_summary', {}).get('period', 'N/A')
        elements.append(Paragraph(
            f"Prepared for: <b>{user.username}</b> ({user.email}) &nbsp;|&nbsp; Period: <b>{period_str}</b> &nbsp;|&nbsp; Statement Date: <b>{timezone.now().strftime('%B %d, %Y')}</b>",
            subtitle_style
        ))
        elements.append(Spacer(1, 10))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#2563EB'), spaceAfter=14))

        # 1. Financial Summary KPI Section
        elements.append(Paragraph("1. Financial Summary Overview", section_heading))
        fin = report_data.get('financial_summary', {})
        
        inc_val = fin.get('total_income', 0.0)
        exp_val = fin.get('total_expense', 0.0)
        bal_val = fin.get('current_balance', 0.0)
        bgt_val = fin.get('total_budget', 0.0)
        rem_bgt_val = fin.get('remaining_budget', 0.0)
        sav_val = fin.get('total_savings', 0.0)

        fin_table_data = [
            [
                Paragraph('Total Income', cell_bold_left), Paragraph(f"+${inc_val:,.2f}", cell_bold_right),
                Paragraph('Total Expense', cell_bold_left), Paragraph(f"-${exp_val:,.2f}", cell_bold_right)
            ],
            [
                Paragraph('Current Balance', cell_bold_left), Paragraph(f"${bal_val:,.2f}", cell_bold_right),
                Paragraph('Total Budget', cell_bold_left), Paragraph(f"${bgt_val:,.2f}", cell_bold_right)
            ],
            [
                Paragraph('Remaining Budget', cell_bold_left), Paragraph(f"${rem_bgt_val:,.2f}", cell_bold_right),
                Paragraph('Total Savings', cell_bold_left), Paragraph(f"${sav_val:,.2f}", cell_bold_right)
            ]
        ]

        fin_table = Table(fin_table_data, colWidths=[130, 140, 130, 140])
        fin_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(fin_table)
        elements.append(Spacer(1, 14))

        # 2. Expense History Section
        elements.append(Paragraph("2. Expense History Breakdown", section_heading))
        expenses = report_data.get('expense_summary', {}).get('expenses', [])
        exp_table_data = [[
            Paragraph('Date', cell_hdr), Paragraph('Expense Title', cell_hdr),
            Paragraph('Category', cell_hdr), Paragraph('Amount', cell_hdr_right),
            Paragraph('Description', cell_hdr)
        ]]

        for i, item in enumerate(expenses[:30]):
            amt = float(item.get('amount', 0))
            is_even = (i % 2 == 0)
            exp_table_data.append([
                Paragraph(str(item.get('date', '')), cell_left),
                Paragraph(str(item.get('title', '')), cell_bold_left),
                Paragraph(str(item.get('category', '')), cell_left),
                Paragraph(f"-${amt:,.2f}", cell_bold_right),
                Paragraph(str(item.get('description') or '—'), cell_left)
            ])

        if len(exp_table_data) == 1:
            exp_table_data.append([Paragraph('No expenses recorded', cell_left), '', '', '', ''])

        exp_table = Table(exp_table_data, colWidths=[70, 130, 90, 80, 170])
        exp_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#0F172A')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        # Zebra striping
        for r_idx in range(1, len(exp_table_data)):
            if r_idx % 2 == 0:
                exp_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#F1F5F9')))
            else:
                exp_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.white))

        exp_table.setStyle(TableStyle(exp_style))
        elements.append(exp_table)
        elements.append(Spacer(1, 14))

        # 3. Savings Goal Progress Section
        elements.append(Paragraph("3. Savings Goal Progress", section_heading))
        goals = report_data.get('savings_summary', {}).get('goals', [])
        goal_table_data = [[
            Paragraph('Goal Name', cell_hdr), Paragraph('Target', cell_hdr_right),
            Paragraph('Saved', cell_hdr_right), Paragraph('Remaining', cell_hdr_right),
            Paragraph('Progress', cell_hdr_right), Paragraph('Status', cell_hdr)
        ]]

        for i, g in enumerate(goals[:15]):
            target = float(g.get('target_amount', 0))
            saved = float(g.get('saved_amount', 0))
            remaining = float(g.get('remaining_amount', max(0, target - saved)))
            progress = float(g.get('progress_percentage', 0))
            status_str = str(g.get('status') or g.get('goal_status', ''))

            goal_table_data.append([
                Paragraph(str(g.get('goal_name', '')), cell_bold_left),
                Paragraph(f"${target:,.2f}", cell_right),
                Paragraph(f"${saved:,.2f}", cell_bold_right),
                Paragraph(f"${remaining:,.2f}", cell_right),
                Paragraph(f"{progress:.1f}%", cell_bold_right),
                Paragraph(status_str, cell_left)
            ])

        if len(goal_table_data) == 1:
            goal_table_data.append([Paragraph('No savings goals created', cell_left), '', '', '', '', ''])

        goal_table = Table(goal_table_data, colWidths=[130, 80, 80, 80, 80, 90])
        goal_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#0F172A')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        for r_idx in range(1, len(goal_table_data)):
            if r_idx % 2 == 0:
                goal_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#F1F5F9')))
            else:
                goal_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.white))

        goal_table.setStyle(TableStyle(goal_style))
        elements.append(goal_table)

        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_excel(user, report_data):
        wb = openpyxl.Workbook()

        title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        totals_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        totals_font = Font(name="Calibri", size=11, bold=True, color="0F172A")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        double_bottom_border = Border(
            top=Side(style='thin', color='0F172A'),
            bottom=Side(style='double', color='0F172A'),
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1')
        )

        # Priority fills
        high_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        high_font = Font(name="Calibri", size=11, bold=True, color="B91C1C")

        medium_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        medium_font = Font(name="Calibri", size=11, bold=True, color="D97706")

        low_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        low_font = Font(name="Calibri", size=11, bold=True, color="15803D")

        completed_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

        def apply_standard_sheet_styling(ws, title, headers, start_row=4):
            ws.views.sheetView[0].showGridLines = True

            # Title Row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 2))
            t_cell = ws.cell(row=1, column=1, value=title.upper())
            t_cell.fill = title_fill
            t_cell.font = title_font
            t_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40

            # Subtitle
            sub_cell = ws.cell(row=2, column=1, value=f"User: {user.username} ({user.email}) | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
            sub_cell.font = Font(name="Calibri", size=9, italic=True, color="64748B")
            ws.row_dimensions[2].height = 18

            # Header Row
            header_row_idx = start_row
            ws.row_dimensions[header_row_idx].height = 26
            for col_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=header_row_idx, column=col_idx, value=h_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Freeze headers below header row
            ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

        def auto_fit_and_filter(ws, headers, header_row=4, total_rows=0):
            # AutoFilter
            last_row = header_row + total_rows
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(last_row, header_row+1)}"

            # Auto Column Widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

        # ----------------------------------------------------
        # Sheet 1: Financial Summary & KPI Cards
        # ----------------------------------------------------
        ws1 = wb.active
        ws1.title = "Financial Summary"
        fin = report_data.get('financial_summary', {})

        apply_standard_sheet_styling(ws1, "BudgetBuddy — Financial Summary", ["Metric / Category", "Value / Amount"], start_row=4)

        # KPI Cards (Rows 4 to 9 merged layout or colored summary block)
        kpi_items = [
            ("Total Income", fin.get('total_income', 0.0), "DCFCE7", "15803D"),
            ("Total Expense", fin.get('total_expense', 0.0), "FEE2E2", "B91C1C"),
            ("Current Balance", fin.get('current_balance', 0.0), "DBEAFE", "1D4ED8"),
            ("Total Budget", fin.get('total_budget', 0.0), "F3E8FF", "7E22CE"),
            ("Remaining Budget", fin.get('remaining_budget', 0.0), "E0F2FE", "0369A1"),
            ("Total Savings", fin.get('total_savings', 0.0), "FEF3C7", "D97706"),
        ]

        curr_row = 5
        for label, val, bg_hex, fg_hex in kpi_items:
            c1 = ws1.cell(row=curr_row, column=1, value=label)
            c2 = ws1.cell(row=curr_row, column=2, value=val)
            
            c1.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
            c2.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
            
            c1.font = Font(name="Calibri", size=11, bold=True, color=fg_hex)
            c2.font = Font(name="Calibri", size=12, bold=True, color=fg_hex)
            
            c1.border = thin_border
            c2.border = thin_border
            c2.number_format = "$#,##0.00"
            c2.alignment = Alignment(horizontal="right")
            ws1.row_dimensions[curr_row].height = 24
            curr_row += 1

        auto_fit_and_filter(ws1, ["Metric / Category", "Value / Amount"], header_row=4, total_rows=len(kpi_items))

        # ----------------------------------------------------
        # Sheet 2: Income
        # ----------------------------------------------------
        ws2 = wb.create_sheet(title="Income")
        headers2 = ["Date", "Source", "Amount ($)", "Description"]
        apply_standard_sheet_styling(ws2, "Income History", headers2)

        incomes_qs = Income.objects.filter(user=user).order_by('-income_date')
        r_idx = 5
        for inc in incomes_qs:
            row_fill = zebra_fill if (r_idx % 2 == 0) else white_fill
            
            c_date = ws2.cell(row=r_idx, column=1, value=inc.income_date.isoformat())
            c_src = ws2.cell(row=r_idx, column=2, value=inc.source)
            c_amt = ws2.cell(row=r_idx, column=3, value=float(inc.amount))
            c_desc = ws2.cell(row=r_idx, column=4, value=inc.description or '—')

            for c in [c_date, c_src, c_amt, c_desc]:
                c.fill = row_fill
                c.border = thin_border

            c_amt.number_format = "$#,##0.00"
            c_amt.alignment = Alignment(horizontal="right")
            r_idx += 1

        auto_fit_and_filter(ws2, headers2, header_row=4, total_rows=incomes_qs.count())

        # ----------------------------------------------------
        # Sheet 3: Expenses (With Totals & Sort by newest date first)
        # ----------------------------------------------------
        ws3 = wb.create_sheet(title="Expenses")
        headers3 = ["Date", "Expense Title", "Category", "Amount ($)", "Description"]
        apply_standard_sheet_styling(ws3, "Expense History", headers3)

        expenses_qs = Expense.objects.filter(user=user).order_by('-expense_date', '-created_at')
        r_idx = 5
        for exp in expenses_qs:
            row_fill = zebra_fill if (r_idx % 2 == 0) else white_fill

            c_date = ws3.cell(row=r_idx, column=1, value=exp.expense_date.isoformat())
            c_title = ws3.cell(row=r_idx, column=2, value=exp.title)
            c_cat = ws3.cell(row=r_idx, column=3, value=exp.category)
            c_amt = ws3.cell(row=r_idx, column=4, value=float(exp.amount))
            c_desc = ws3.cell(row=r_idx, column=5, value=exp.description or '—')

            for c in [c_date, c_title, c_cat, c_amt, c_desc]:
                c.fill = row_fill
                c.border = thin_border

            c_amt.number_format = "$#,##0.00"
            c_amt.alignment = Alignment(horizontal="right")
            r_idx += 1

        # Add Totals Row at bottom
        if expenses_qs.exists():
            t_lbl = ws3.cell(row=r_idx, column=1, value="TOTAL EXPENSES")
            t_lbl.font = totals_font
            t_lbl.fill = totals_fill
            t_lbl.border = double_bottom_border

            for c_i in [2, 3, 5]:
                c_fill = ws3.cell(row=r_idx, column=c_i, value="")
                c_fill.fill = totals_fill
                c_fill.border = double_bottom_border

            t_sum = ws3.cell(row=r_idx, column=4, value=f"=SUM(D5:D{r_idx-1})")
            t_sum.font = totals_font
            t_sum.fill = totals_fill
            t_sum.border = double_bottom_border
            t_sum.number_format = "$#,##0.00"
            t_sum.alignment = Alignment(horizontal="right")
            ws3.row_dimensions[r_idx].height = 24

        auto_fit_and_filter(ws3, headers3, header_row=4, total_rows=expenses_qs.count() + 1)

        # ----------------------------------------------------
        # Sheet 4: Budgets
        # ----------------------------------------------------
        ws4 = wb.create_sheet(title="Budgets")
        headers4 = ["Category", "Budget Amount ($)", "Month", "Year"]
        apply_standard_sheet_styling(ws4, "Category Budgets", headers4)

        budgets_qs = Budget.objects.filter(user=user).order_by('year', 'month')
        r_idx = 5
        for bgt in budgets_qs:
            row_fill = zebra_fill if (r_idx % 2 == 0) else white_fill

            c_cat = ws4.cell(row=r_idx, column=1, value=bgt.category)
            c_amt = ws4.cell(row=r_idx, column=2, value=float(bgt.budget_amount))
            c_m = ws4.cell(row=r_idx, column=3, value=bgt.month)
            c_y = ws4.cell(row=r_idx, column=4, value=bgt.year)

            for c in [c_cat, c_amt, c_m, c_y]:
                c.fill = row_fill
                c.border = thin_border

            c_amt.number_format = "$#,##0.00"
            c_amt.alignment = Alignment(horizontal="right")
            r_idx += 1

        auto_fit_and_filter(ws4, headers4, header_row=4, total_rows=budgets_qs.count())

        # ----------------------------------------------------
        # Sheet 5: Savings Goals (Highlight Completed Goals in Green)
        # ----------------------------------------------------
        ws5 = wb.create_sheet(title="Savings Goals")
        headers5 = ["Goal Name", "Target Amount ($)", "Saved Amount ($)", "Remaining Amount ($)", "Progress (%)", "Status", "Target Date"]
        apply_standard_sheet_styling(ws5, "Savings Goals Overview", headers5)

        goals_qs = SavingsGoal.objects.filter(user=user).order_by('target_date')
        r_idx = 5
        for g in goals_qs:
            target = float(g.target_amount)
            saved = float(g.saved_amount)
            remaining = max(0.0, target - saved)
            progress = (saved / target) if target > 0 else 0.0

            is_completed = (g.status == 'COMPLETED' or saved >= target)
            row_fill = completed_fill if is_completed else (zebra_fill if (r_idx % 2 == 0) else white_fill)

            c_name = ws5.cell(row=r_idx, column=1, value=g.goal_name)
            c_target = ws5.cell(row=r_idx, column=2, value=target)
            c_saved = ws5.cell(row=r_idx, column=3, value=saved)
            c_rem = ws5.cell(row=r_idx, column=4, value=remaining)
            c_pct = ws5.cell(row=r_idx, column=5, value=progress)
            c_stat = ws5.cell(row=r_idx, column=6, value=g.status)
            c_date = ws5.cell(row=r_idx, column=7, value=g.target_date.isoformat())

            for c in [c_name, c_target, c_saved, c_rem, c_pct, c_stat, c_date]:
                c.fill = row_fill
                c.border = thin_border

            c_target.number_format = "$#,##0.00"
            c_saved.number_format = "$#,##0.00"
            c_rem.number_format = "$#,##0.00"
            c_pct.number_format = "0.0%"

            c_target.alignment = Alignment(horizontal="right")
            c_saved.alignment = Alignment(horizontal="right")
            c_rem.alignment = Alignment(horizontal="right")
            c_pct.alignment = Alignment(horizontal="right")

            if is_completed:
                c_stat.font = Font(name="Calibri", size=11, bold=True, color="15803D")

            r_idx += 1

        auto_fit_and_filter(ws5, headers5, header_row=4, total_rows=goals_qs.count())

        # ----------------------------------------------------
        # Sheet 6: Notifications (Color Priority Badges: HIGH=Red, MEDIUM=Orange, LOW=Green)
        # ----------------------------------------------------
        ws6 = wb.create_sheet(title="Notifications")
        headers6 = ["Date & Time", "Priority", "Notification Type", "Title", "Message"]
        apply_standard_sheet_styling(ws6, "User Notifications History", headers6)

        notifs_qs = Notification.objects.filter(user=user).order_by('-created_at')
        r_idx = 5
        for n in notifs_qs:
            row_fill = zebra_fill if (r_idx % 2 == 0) else white_fill

            c_date = ws6.cell(row=r_idx, column=1, value=n.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            c_prio = ws6.cell(row=r_idx, column=2, value=n.priority)
            c_type = ws6.cell(row=r_idx, column=3, value=n.notification_type)
            c_title = ws6.cell(row=r_idx, column=4, value=n.title)
            c_msg = ws6.cell(row=r_idx, column=5, value=n.message)

            for c in [c_date, c_prio, c_type, c_title, c_msg]:
                c.fill = row_fill
                c.border = thin_border

            # Color priority badges
            p_val = (n.priority or '').upper()
            if p_val == 'HIGH':
                c_prio.fill = high_fill
                c_prio.font = high_font
            elif p_val == 'MEDIUM':
                c_prio.fill = medium_fill
                c_prio.font = medium_font
            elif p_val == 'LOW':
                c_prio.fill = low_fill
                c_prio.font = low_font

            c_prio.alignment = Alignment(horizontal="center")
            r_idx += 1

        auto_fit_and_filter(ws6, headers6, header_row=4, total_rows=notifs_qs.count())

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

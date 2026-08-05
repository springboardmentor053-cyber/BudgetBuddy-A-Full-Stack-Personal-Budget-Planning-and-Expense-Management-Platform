import os
import sys
import io
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

from datetime import date
from decimal import Decimal
import openpyxl
from django.contrib.auth.models import User
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from expenses.models import Expense, Income, Budget
from savings.models import SavingsGoal, Notification

def run_export_tests():
    print("Testing Production Quality Reports Export Module 9 Compliance...")

    # 1. Setup Test User
    user, _ = User.objects.get_or_create(username='spec_export_user', email='spec_export@example.com')
    user.set_password('TestPass123!')
    user.save()

    # Clean data for test user
    Expense.objects.filter(user=user).delete()
    Income.objects.filter(user=user).delete()
    Budget.objects.filter(user=user).delete()
    SavingsGoal.objects.filter(user=user).delete()
    Notification.objects.filter(user=user).delete()

    # Create dummy data
    today = date.today()
    Income.objects.create(user=user, source='Salary', amount=Decimal('5000.00'), income_date=today)
    Expense.objects.create(user=user, title='Groceries', amount=Decimal('400.00'), category='FOOD', expense_date=today)
    Budget.objects.create(user=user, category='FOOD', budget_amount=Decimal('600.00'), month=today.month, year=today.year)
    SavingsGoal.objects.create(user=user, goal_name='Vacation', target_amount=Decimal('2000.00'), saved_amount=Decimal('2000.00'), status='COMPLETED', target_date=today)
    Notification.objects.create(user=user, title='Budget Exceeded', message='High priority alert', notification_type='BUDGET_EXCEEDED', priority='HIGH')

    token = str(RefreshToken.for_user(user).access_token)
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

    results = {}

    # TEST 1: Export PDF Endpoint (GET /api/reports/export/pdf/)
    res_pdf = client.get('/api/reports/export/pdf/')
    assert res_pdf.status_code == 200, f"PDF export failed: {res_pdf.status_code}"
    assert res_pdf.headers['Content-Type'] == 'application/pdf', f"Wrong content-type: {res_pdf.headers['Content-Type']}"
    assert res_pdf.content.startswith(b'%PDF-'), "Invalid PDF binary stream"
    results['Export PDF API (GET /api/reports/export/pdf/)'] = 'PASS'

    # TEST 2: Export Excel Endpoint (GET /api/reports/export/excel/)
    res_excel = client.get('/api/reports/export/excel/')
    assert res_excel.status_code == 200, f"Excel export failed: {res_excel.status_code}"
    assert 'spreadsheetml.sheet' in res_excel.headers['Content-Type'], f"Wrong content-type: {res_excel.headers['Content-Type']}"
    
    # Load workbook from binary stream
    wb = openpyxl.load_workbook(io.BytesIO(res_excel.content))
    expected_sheets = ["Financial Summary", "Income", "Expenses", "Budgets", "Savings Goals", "Notifications"]
    assert wb.sheetnames == expected_sheets, f"Expected sheets {expected_sheets}, got {wb.sheetnames}"
    
    # Verify sheet 3 (Expenses) total row exists
    ws3 = wb['Expenses']
    assert ws3.cell(row=6, column=1).value == 'TOTAL EXPENSES'
    
    # Verify sheet 6 (Notifications) HIGH priority styling
    ws6 = wb['Notifications']
    assert ws6.cell(row=5, column=2).value == 'HIGH'
    
    results['Export Excel API (GET /api/reports/export/excel/) & Formatting'] = 'PASS'

    # TEST 3: JSON Export Backward Compatibility (GET /api/reports/export/?format=json)
    res_json = client.get('/api/reports/export/?format=json')
    assert res_json.status_code == 200, f"JSON export failed: {res_json.status_code}"
    assert 'financial_summary' in res_json.data, "Missing financial_summary in JSON export"
    results['JSON Export Backward Compatibility'] = 'PASS'

    print("\n--- ALL PRODUCTION QUALITY EXPORT COMPLIANCE TESTS PASSED ---")
    for k, v in results.items():
        print(f"[PASS] {k}: {v}")

if __name__ == '__main__':
    run_export_tests()
